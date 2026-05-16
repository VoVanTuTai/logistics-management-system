import openpyxl
import sys

def read_xlsx(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                print("\t".join([str(cellValue) if cellValue is not None else "" for cellValue in row]))
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        read_xlsx(sys.argv[1])
    else:
        print("Usage: python read_xlsx.py <file_path>")
